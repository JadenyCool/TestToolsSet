import openpyxl


def ReadExcel(path):
    excel_row_start = 0
    excelStart = int(excel_row_start) - 1
    openfile = openpyxl.load_workbook(filename=path, data_only=True)
    getSheetName = openfile.get_sheet_by_name("String")

    excelChineseList = []

    # 读取excel中的中文
    excelfuhao = ["。", ":", "：", ".", ";"]
    aa = {}

    # Ckey_value = getSheetName.cell(row=1, column=6).value
    # print(Ckey_value)
    # 获取单元格值：
    # Data = table.cell(row=row, column=col).value
    j = 1
    for excel_translate in list(getSheetName.columns)[5]:
        excel_value = excel_translate.value  # 小语种

        # print(Ckey_value)
        if excel_value is None:
           j = j+1
        else:
            Ckey_value = getSheetName.cell(row=j, column=1).value
            for i in range(len(excelfuhao)):
                if Ckey_value.endswith("{}".format(excelfuhao[i])):
                    Ckey_value = Ckey_value.replace("{}".format(excelfuhao[i]), "")
                    break
                if excel_value.endswith("{}".format(excelfuhao[i])):
                    excel_value = excel_value.replace("{}".format(excelfuhao[i]), "")
                    break
            aa.update({Ckey_value: excel_value})
            j = j+1

    for k, v in aa.items():
        print("{}------------{}".format(k, v))


if __name__ == '__main__':
    filepath = "D:/ddd/7tran/tranOne/tranOne.xlsx"
    ReadExcel(filepath)
